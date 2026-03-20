"""DeepAgents 统一智能体服务 - v3.0 生产环境重构

核心设计原则：
- 智能体专注于回答用户问题
- 提供细粒度的 Schema 查询工具，实现渐进式上下文加载
- 使用 LangGraph checkpointer 实现状态持久化
- 使用原生异步 API，避免线程与协程混合

重构内容：
- 工具定义使用 @tool 装饰器和 InjectedToolArg
- 使用 session_id 作为 thread_id
- 配置 checkpointer 和 store
- 使用异步流式 API 替代 threading
"""

import json
import re
import uuid
from datetime import datetime, timedelta
from typing import Any, AsyncGenerator, Optional, List
from dataclasses import dataclass, asdict
import asyncio
from contextvars import ContextVar

from langchain_core.tools import tool

from app.core import get_settings

settings = get_settings()


# ==================== 请求上下文（传递数据库连接参数） ====================

# 当前请求的数据库连接上下文
_db_context: ContextVar[dict] = ContextVar('db_context')


def set_db_context(
    datasource_id: str,
    db_type: str,
    host: str = None,
    port: int = None,
    database: str = None,
    username: str = None,
    password: str = None,
    schema_name: Optional[str] = None,
    tables_info: Optional[List[dict]] = None,
):
    """设置当前请求的数据库连接上下文"""
    _db_context.set({
        "datasource_id": datasource_id,
        "db_type": db_type,
        "host": host,
        "port": port,
        "database": database,
        "username": username,
        "password": password,
        "schema_name": schema_name,
        "tables_info": tables_info,
    })


def get_db_context() -> Optional[dict]:
    """获取当前请求的数据库连接上下文"""
    try:
        return _db_context.get()
    except LookupError:
        return None


# ==================== SSE 事件类型定义 ====================

@dataclass
class SSEEvent:
    """SSE 事件基类"""
    type: str
    timestamp: str = None

    def __post_init__(self):
        if self.timestamp is None:
            self.timestamp = datetime.utcnow().isoformat()

    def to_dict(self) -> dict:
        return {k: v for k, v in asdict(self).items() if v is not None}


@dataclass
class ThinkingEvent(SSEEvent):
    """思考事件"""
    type: str = "thinking"
    content: str = None


@dataclass
class PlanEvent(SSEEvent):
    """计划事件"""
    type: str = "plan"
    todos: List[dict] = None


@dataclass
class ToolCallEvent(SSEEvent):
    """工具调用事件"""
    type: str = "tool_call"
    name: str = None  # 显示名称，如 "列出数据表"
    tool: str = None
    input: dict = None
    id: str = None


@dataclass
class ToolResultEvent(SSEEvent):
    """工具结果事件"""
    type: str = "tool_result"
    name: str = None  # 显示名称
    tool: str = None
    id: str = None
    output: str = None


@dataclass
class SQLGenerationEvent(SSEEvent):
    """SQL 生成事件"""
    type: str = "sql_generation"
    name: str = "生成 SQL"
    sql: str = None


@dataclass
class SQLExecutionEvent(SSEEvent):
    """SQL 执行事件"""
    type: str = "sql_execution"
    name: str = "执行查询"
    status: str = None
    duration: float = None


@dataclass
class MessageEvent(SSEEvent):
    """消息事件"""
    type: str = "message"
    content: str = None


@dataclass
class ErrorEvent(SSEEvent):
    """错误事件"""
    type: str = "error"
    message: str = None
    code: str = None


@dataclass
class DoneEvent(SSEEvent):
    """完成事件"""
    type: str = "done"
    message: str = "处理完成"


# ==================== 系统提示词 ====================

SYSTEM_PROMPT = """你是 SmartNum 数据分析助手，帮助用户查询和分析数据库中的数据。

## 工具

### explore_query_guide
使用 shell 命令浏览**当前数据源**的查询指南文档（ls, cat, grep 等）。
查询指南是一个文件夹，包含多个文档（文档内容可能涵盖业务说明、统计口径、表字段说明等），命令会自动在当前数据源的查询指南文件夹下执行。

### list_tables
列出数据库中的表。

### get_table_schema
获取表的字段结构。

### run_sql
执行 SELECT 查询。

### render_chart / export_data
图表渲染和数据导出。

## 输出

数据用 Markdown 表格展示。只执行 SELECT 语句，不查询敏感数据。
"""


# ==================== 核心工具定义 ====================

@tool
async def list_tables(
    datasource_id: str,
) -> str:
    """列出数据库中所有表的名称和注释。

    在查询数据前，先调用此工具了解有哪些表可用。
    返回简洁的表列表，不包含列信息。

    Args:
        datasource_id: 数据源 ID

    Returns:
        表名和注释列表（Markdown 格式）
    """
    from app.services import db_service

    # 从上下文获取数据库连接参数
    ctx = get_db_context()
    if ctx is None:
        return "错误：未找到数据库连接上下文，请确保在正确的会话中调用"

    # 文件类型使用 DuckDB
    if ctx["db_type"] == "file":
        from app.services.file_datasource_service import FileDatasourceService
        service = FileDatasourceService()
        tables_info = ctx.get("tables_info") or await service.get_tables_info(ctx["datasource_id"])

        lines = ["# 数据表列表\n"]
        lines.append("| 序号 | 表名 | 行数 |")
        lines.append("|------|------|------|")

        for i, table in enumerate(tables_info, 1):
            table_name = table.get("name", "unknown")
            row_count = table.get("row_count", 0)
            lines.append(f"| {i} | {table_name} | {row_count} |")

        return "\n".join(lines)

    # 数据库类型使用原有逻辑
    url = db_service.get_database_url(
        ctx["db_type"],
        ctx["host"],
        ctx["port"],
        ctx["database"],
        ctx["username"],
        ctx["password"],
    )

    from sqlalchemy import text
    from sqlalchemy.ext.asyncio import create_async_engine

    engine = create_async_engine(url, echo=False)

    try:
        async with engine.connect() as conn:
            lines = ["# 数据库表列表\n"]
            lines.append("| 序号 | 表名 | 说明 |")
            lines.append("|------|------|------|")

            if ctx["db_type"] == "mysql":
                result = await conn.execute(
                    text("""
                        SELECT TABLE_NAME, TABLE_COMMENT
                        FROM information_schema.TABLES
                        WHERE TABLE_SCHEMA = :db
                        ORDER BY TABLE_NAME
                    """),
                    {"db": ctx["database"]},
                )
                rows = result.fetchall()

                for i, row in enumerate(rows, 1):
                    table_name = row[0]
                    table_comment = row[1] or "-"
                    lines.append(f"| {i} | {table_name} | {table_comment} |")

            elif ctx["db_type"] == "postgresql":
                schema = ctx.get("schema_name") or "public"
                result = await conn.execute(
                    text("""
                        SELECT table_name, obj_description((table_schema || '.' || table_name)::regclass) as table_comment
                        FROM information_schema.tables
                        WHERE table_schema = :schema AND table_type = 'BASE TABLE'
                        ORDER BY table_name
                    """),
                    {"schema": schema},
                )
                rows = result.fetchall()

                for i, row in enumerate(rows, 1):
                    table_name = row[0]
                    table_comment = row[1] or "-"
                    lines.append(f"| {i} | {table_name} | {table_comment} |")

            elif ctx["db_type"] == "sqlite":
                result = await conn.execute(
                    text("SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%' ORDER BY name")
                )
                rows = result.fetchall()

                for i, row in enumerate(rows, 1):
                    table_name = row[0]
                    lines.append(f"| {i} | {table_name} | - |")

            return "\n".join(lines)

    finally:
        await engine.dispose()


@tool
async def get_table_schema(
    datasource_id: str,
    table_name: str,
) -> str:
    """获取指定表的详细结构信息。

    当你需要了解某个表的列名、类型、注释等信息时调用此工具。
    建议只查询与用户问题相关的表，避免加载过多信息。

    Args:
        datasource_id: 数据源 ID
        table_name: 表名

    Returns:
        表结构信息（Markdown 格式），包含列名、类型、注释等
    """
    from app.services import db_service

    # 从上下文获取数据库连接参数
    ctx = get_db_context()
    if ctx is None:
        return "错误：未找到数据库连接上下文，请确保在正确的会话中调用"

    # 文件类型使用 DuckDB
    if ctx["db_type"] == "file":
        from app.services.file_datasource_service import FileDatasourceService
        service = FileDatasourceService()
        tables_info = ctx.get("tables_info") or await service.get_tables_info(ctx["datasource_id"])

        # 查找指定表
        table_info = None
        for t in tables_info:
            if t.get("name") == table_name:
                table_info = t
                break

        if not table_info:
            return f"错误：表 '{table_name}' 不存在"

        lines = [f"# 表结构：{table_name}\n"]
        lines.append(f"**行数**: {table_info.get('row_count', 0)}\n")
        lines.append("\n| 列名 | 类型 | 可空 |")
        lines.append("|------|------|------|")

        for col in table_info.get("columns", []):
            col_name = col.get("name", "unknown")
            col_type = col.get("type", "unknown")
            nullable = "是" if col.get("nullable", True) else "否"
            lines.append(f"| {col_name} | {col_type} | {nullable} |")

        return "\n".join(lines)

    # 数据库类型使用原有逻辑
    url = db_service.get_database_url(
        ctx["db_type"],
        ctx["host"],
        ctx["port"],
        ctx["database"],
        ctx["username"],
        ctx["password"],
    )

    from sqlalchemy import text
    from sqlalchemy.ext.asyncio import create_async_engine

    engine = create_async_engine(url, echo=False)

    try:
        async with engine.connect() as conn:
            lines = [f"# 表结构：{table_name}\n"]

            if ctx["db_type"] == "mysql":
                # 获取表注释
                table_result = await conn.execute(
                    text("""
                        SELECT TABLE_COMMENT
                        FROM information_schema.TABLES
                        WHERE TABLE_SCHEMA = :db AND TABLE_NAME = :table
                    """),
                    {"db": ctx["database"], "table": table_name},
                )
                table_row = table_result.fetchone()
                if table_row and table_row[0]:
                    lines.append(f"**表说明**: {table_row[0]}\n")

                # 获取列信息
                columns_result = await conn.execute(
                    text("""
                        SELECT
                            COLUMN_NAME,
                            COLUMN_TYPE,
                            IS_NULLABLE,
                            COLUMN_KEY,
                            COLUMN_DEFAULT,
                            COLUMN_COMMENT
                        FROM information_schema.COLUMNS
                        WHERE TABLE_SCHEMA = :db AND TABLE_NAME = :table
                        ORDER BY ORDINAL_POSITION
                    """),
                    {"db": ctx["database"], "table": table_name},
                )
                column_rows = columns_result.fetchall()

                lines.append("\n| 列名 | 类型 | 可空 | 键 | 默认值 | 说明 |")
                lines.append("|------|------|------|-----|--------|------|")

                for col in column_rows:
                    col_name = col[0]
                    col_type = col[1]
                    nullable = "是" if col[2] == "YES" else "否"
                    col_key = col[3] or "-"
                    col_default = col[4] or "-"
                    col_comment = col[5] or "-"
                    lines.append(f"| {col_name} | {col_type} | {nullable} | {col_key} | {col_default} | {col_comment} |")

            elif ctx["db_type"] == "postgresql":
                schema = ctx.get("schema_name") or "public"

                # 获取列信息
                columns_result = await conn.execute(
                    text("""
                        SELECT
                            column_name,
                            data_type || COALESCE('(' || character_maximum_length || ')', ''),
                            is_nullable,
                            NULL as column_key,
                            column_default,
                            col_description((table_schema || '.' || table_name)::regclass, ordinal_position)
                        FROM information_schema.columns
                        WHERE table_schema = :schema AND table_name = :table
                        ORDER BY ordinal_position
                    """),
                    {"schema": schema, "table": table_name},
                )
                column_rows = columns_result.fetchall()

                lines.append("\n| 列名 | 类型 | 可空 | 键 | 默认值 | 说明 |")
                lines.append("|------|------|------|-----|--------|------|")

                for col in column_rows:
                    col_name = col[0]
                    col_type = col[1]
                    nullable = "是" if col[2] == "YES" else "否"
                    col_key = col[3] or "-"
                    col_default = col[4] or "-"
                    col_comment = col[5] or "-"
                    lines.append(f"| {col_name} | {col_type} | {nullable} | {col_key} | {col_default} | {col_comment} |")

            elif ctx["db_type"] == "sqlite":
                # 获取表结构
                pragma_result = await conn.execute(
                    text(f"PRAGMA table_info({table_name})")
                )
                column_rows = pragma_result.fetchall()

                lines.append("\n| 列名 | 类型 | 可空 | 主键 | 默认值 |")
                lines.append("|------|------|------|------|--------|")

                for col in column_rows:
                    col_name = col[1]
                    col_type = col[2] or "-"
                    nullable = "否" if col[3] == 1 else "是"
                    is_pk = "是" if col[5] == 1 else "否"
                    col_default = col[4] or "-"
                    lines.append(f"| {col_name} | {col_type} | {nullable} | {is_pk} | {col_default} |")

            return "\n".join(lines)

    except Exception as e:
        return f"错误：获取表结构失败 - {str(e)}"

    finally:
        await engine.dispose()


@tool
async def run_sql(
    datasource_id: str,
    sql: str,
    limit: int = 1000,
) -> dict:
    """执行 SQL 查询。

    Args:
        datasource_id: 数据源 ID
        sql: SQL 查询语句（仅支持 SELECT）
        limit: 最大返回行数

    Returns:
        查询结果，包含 success、columns、rows、total、truncated、error
    """
    from app.services import db_service

    # 从上下文获取数据库连接参数
    ctx = get_db_context()
    if ctx is None:
        return {"success": False, "error": "未找到数据库连接上下文"}

    # 文件类型使用 DuckDB
    if ctx["db_type"] == "file":
        from app.services.file_datasource_service import FileDatasourceService
        service = FileDatasourceService()
        result = await service.execute_query(
            datasource_id=ctx["datasource_id"],
            sql=sql,
            max_rows=limit,
        )
        return result

    # 数据库类型使用原有逻辑
    result = await db_service.execute_query(
        db_type=ctx["db_type"],
        host=ctx["host"],
        port=ctx["port"],
        database=ctx["database"],
        username=ctx["username"],
        password=ctx["password"],
        sql=sql,
        max_rows=limit,
    )
    return result


@tool
def render_chart(
    chart_type: str,
    title: str,
    data: list,
    x_field: str = None,
    y_field: str = None,
) -> dict:
    """生成 ECharts 图表配置。

    当用户要求可视化时调用此工具。智能体根据查询结果选择合适的图表类型，
    并指定数据字段映射，工具会生成完整的 ECharts 配置选项。

    Args:
        chart_type: 图表类型 - bar (柱状图), line (折线图), pie (饼图), scatter (散点图), area (面积图)
        title: 图表标题
        data: 数据数组，包含图表需要展示的所有数据
        x_field: X 轴字段名（饼图不需要）
        y_field: Y 轴字段名（饼图为数值字段，其他为数值字段）

    Returns:
        ECharts 配置对象，可直接传递给前端的 ECharts 组件

    Examples:
        # 柱状图示例
        render_chart(
            chart_type="bar",
            title="销售额前 10 产品",
            data=[
                {"product": "产品 A", "sales": 1000},
                {"product": "产品 B", "sales": 800},
            ],
            x_field="product",
            y_field="sales"
        )

        # 饼图示例
        render_chart(
            chart_type="pie",
            title="销售占比",
            data=[
                {"category": "类别 A", "value": 400},
                {"category": "类别 B", "value": 300},
            ],
            y_field="value"
        )
    """
    # 图表类型映射
    chart_type_map = {
        "bar": "bar",
        "line": "line",
        "pie": "pie",
        "scatter": "scatter",
        "area": "line",
    }

    series_name = y_field or "数值"
    x_data = []
    y_data = []

    for item in data:
        if x_field:
            x_data.append(item.get(x_field, ""))
        y_data.append(item.get(y_field or series_name, 0))

    # 生成 ECharts 配置
    if chart_type == "pie":
        # 如果没有指定 x_field，自动推断（取第一个非 y_field 的字段）
        if not x_field and data:
            for key in data[0].keys():
                if key != y_field:
                    x_field = key
                    break

        pie_data = []
        for i, item in enumerate(data):
            pie_data.append({
                "name": item.get(x_field) if x_field else f"项{i+1}",
                "value": item.get(y_field or series_name, 0)
            })

        option = {
            "title": {"text": title, "left": "center"},
            "tooltip": {
                "trigger": "item",
                "formatter": "{b}: {c} ({d}%)"
            },
            "legend": {
                "orient": "vertical",
                "left": "left"
            },
            "series": [{
                "name": title,
                "type": "pie",
                "radius": "60%",
                "data": pie_data,
                "emphasis": {
                    "itemStyle": {
                        "shadowBlur": 10,
                        "shadowOffsetX": 0,
                        "shadowColor": "rgba(0, 0, 0, 0.5)"
                    }
                }
            }]
        }
    else:
        option = {
            "title": {"text": title},
            "tooltip": {
                "trigger": "axis",
                "axisPointer": {"type": "shadow"}
            },
            "xAxis": {
                "type": "category",
                "data": x_data,
                "axisLabel": {"rotate": 0 if len(x_data) <= 10 else 45}
            },
            "yAxis": {"type": "value"},
            "series": [{
                "name": series_name,
                "type": chart_type_map.get(chart_type, "bar"),
                "data": y_data,
                **({"areaStyle": {"opacity": 0.3}} if chart_type == "area" else {})
            }]
        }

        if len(x_data) > 10:
            option["grid"] = {"bottom": "15%"}

    return {"chart_type": chart_type, "title": title, "option": option}


@tool
async def export_data(
    filename: str,
    data: list,
    format: str = "csv",
) -> dict:
    """导出数据为文件。

    当用户要求导出表格数据时调用此工具。智能体根据查询结果生成文件，
    工具会返回文件信息和下载标识，前端可通过 download_id 下载文件。

    Args:
        filename: 文件名（不含扩展名）
        data: 数据数组，包含导出的所有数据，每项为一个对象
        format: 导出格式 - csv 或 xlsx

    Returns:
        文件信息对象，包含 filename, format, download_id, size 等

    Examples:
        # 导出 CSV
        export_data(
            filename="产品销售数据",
            data=[
                {"product": "产品 A", "sales": 1000, "count": 50},
                {"product": "产品 B", "sales": 800, "count": 30},
            ],
            format="csv"
        )

        # 导出 Excel
        export_data(
            filename="月度报表",
            data=[
                {"month": "1 月", "revenue": 100000},
                {"month": "2 月", "revenue": 120000},
            ],
            format="xlsx"
        )
    """
    import csv
    import io

    if not data:
        return {"error": "没有可导出的数据"}

    # 生成唯一的下载 ID
    download_id = str(uuid.uuid4())

    # 获取列名
    columns = list(data[0].keys())

    # 生成内容
    if format == "csv":
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=columns, quoting=csv.QUOTE_MINIMAL)
        writer.writeheader()
        writer.writerows(data)
        content = output.getvalue().encode("utf-8-sig")
        file_extension = "csv"
        mime_type = "text/csv"
    elif format == "xlsx":
        try:
            import openpyxl
            from openpyxl import Workbook
        except ImportError:
            return {"error": "Excel 导出需要安装 openpyxl 库"}

        wb = Workbook()
        ws = wb.active
        ws.title = "数据"

        # 写入表头
        for col_idx, col_name in enumerate(columns, 1):
            ws.cell(row=1, column=col_idx, value=col_name)

        # 写入数据
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, col_name in enumerate(columns, 1):
                value = row_data.get(col_name, "")
                # 处理特殊值
                if isinstance(value, datetime):
                    value = value.strftime("%Y-%m-%d %H:%M:%S")
                ws.cell(row=row_idx, column=col_idx, value=value)

        # 自动调整列宽
        for col_idx, col_name in enumerate(columns, 1):
            max_length = len(str(col_name))
            for row_data in data:
                value = row_data.get(col_name, "")
                if value is not None:
                    max_length = max(max_length, len(str(value)))
            col_letter = chr(64 + col_idx) if col_idx <= 26 else f"{chr(64 + col_idx // 26)}{chr(64 + col_idx % 26)}"
            ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

        # 保存到字节流
        output = io.BytesIO()
        wb.save(output)
        content = output.getvalue()
        file_extension = "xlsx"
        mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    else:
        return {"error": f"不支持的导出格式：{format}"}

    # 计算文件大小（KB）
    size_kb = round(len(content) / 1024, 2)

    # 持久化到数据库
    await _save_export_file(
        download_id=download_id,
        filename=f"{filename}.{file_extension}",
        content=content,
        mime_type=mime_type,
        size_kb=size_kb,
    )

    return {
        "download_id": download_id,
        "filename": f"{filename}.{file_extension}",
        "format": format,
        "size": size_kb,
        "row_count": len(data),
        "column_count": len(columns),
    }


# ==================== 导出文件持久化 ====================

async def _save_export_file(
    download_id: str,
    filename: str,
    content: bytes,
    mime_type: str,
    size_kb: int,
    expires_hours: int = 24,
) -> None:
    """保存导出文件到数据库

    Args:
        download_id: 下载 ID
        filename: 文件名
        content: 文件内容
        mime_type: MIME 类型
        size_kb: 文件大小（KB）
        expires_hours: 过期时间（小时）
    """
    from app.models.database import async_session_maker
    from app.models.models import ExportFile

    async with async_session_maker() as session:
        export_file = ExportFile(
            id=download_id,
            filename=filename,
            content=content,
            mime_type=mime_type,
            size_kb=size_kb,
            created_at=datetime.utcnow(),
            expires_at=datetime.utcnow() + timedelta(hours=expires_hours),
        )
        session.add(export_file)
        await session.commit()


async def get_export_file(download_id: str) -> dict | None:
    """根据 download_id 获取文件内容

    Args:
        download_id: 下载 ID

    Returns:
        文件信息字典，包含 content、filename、mime_type 等
    """
    from app.models.database import async_session_maker
    from app.models.models import ExportFile
    from sqlalchemy import select

    async with async_session_maker() as session:
        result = await session.execute(
            select(ExportFile).where(ExportFile.id == download_id)
        )
        export_file = result.scalar_one_or_none()

        if export_file is None:
            return None

        # 检查是否过期
        if export_file.expires_at < datetime.utcnow():
            await session.delete(export_file)
            await session.commit()
            return None

        return {
            "content": export_file.content,
            "filename": export_file.filename,
            "mime_type": export_file.mime_type,
            "created_at": export_file.created_at,
        }


async def cleanup_expired_export_files() -> int:
    """清理过期的导出文件

    Returns:
        删除的文件数量
    """
    from app.models.database import async_session_maker
    from app.models.models import ExportFile
    from sqlalchemy import delete

    async with async_session_maker() as session:
        result = await session.execute(
            delete(ExportFile).where(ExportFile.expires_at < datetime.utcnow())
        )
        await session.commit()
        return result.rowcount


# ==================== 查询指南工具 ====================

@tool
async def explore_query_guide(
    command: str,
) -> str:
    """使用 shell 命令查阅当前数据源的查询指南文件夹。

    查询指南是一个文件夹，里面包含多个文档文件：
    - notes.md - 人工编写的备注说明
    - uploaded/* - 上传的参考文档

    你可以像探索普通文件夹一样，使用 ls 查看有哪些文件，用 cat 阅读内容，用 grep 搜索关键词等。

    Args:
        command: 要执行的 shell 命令（自动在当前数据源的查询指南文件夹下执行）

    Returns:
        命令执行结果

    Examples:
        explore_query_guide("ls -la")  # 查看查询指南文件夹中有哪些文档
        explore_query_guide("ls uploaded/")  # 查看上传了哪些文档
        explore_query_guide("cat notes.md")  # 阅读备注说明
        explore_query_guide("grep -r '活跃用户' .")  # 搜索所有文档中关于"活跃用户"的定义
        explore_query_guide("head -30 uploaded/业务说明.md")  # 查看某个文档的前 30 行
    """
    from app.services.query_guide_service import QueryGuideService

    ctx = get_db_context()
    if ctx is None:
        return "错误：未找到数据库连接上下文"

    datasource_id = ctx.get("datasource_id")
    if not datasource_id:
        return "错误：未找到数据源 ID"

    service = QueryGuideService()
    output = service.explore_guide(
        datasource_id=datasource_id,
        command=command,
    )
    return output


# ==================== DeepAgent 创建 ====================

_agent = None
_checkpointer = None
_store = None


async def get_agent():
    """获取 DeepAgent 实例（带状态持久化）"""
    global _agent, _checkpointer, _store

    if _agent is not None:
        return _agent

    from deepagents import create_deep_agent
    from langchain_openai import ChatOpenAI
    from app.services.checkpointer import get_checkpointer, get_store

    # 获取 checkpointer 和 store
    _checkpointer = await get_checkpointer()
    _store = get_store()

    # 创建 OpenAI 兼容的 LLM
    llm = ChatOpenAI(
        model=settings.llm_model_name,
        api_key=settings.llm_api_key,
        base_url=settings.llm_base_url,
        temperature=settings.llm_temperature,
        max_tokens=settings.llm_max_tokens,
        request_timeout=settings.llm_timeout,
    )

    # 创建智能体 - 添加 checkpointer 和 store
    _agent = create_deep_agent(
        name="smartnum-agent",
        model=llm,
        tools=[
            list_tables, get_table_schema, run_sql, render_chart, export_data,
            explore_query_guide,
        ],
        system_prompt=SYSTEM_PROMPT,
        checkpointer=_checkpointer,
        store=_store,
    )

    return _agent


# ==================== 标题生成 ====================

def generate_session_title(user_message: str) -> str:
    """
    根据用户第一条消息生成会话标题
    直接用户户问题作为标题，去除多余内容
    """
    # 直接用户户问题作为标题，限制长度
    title = user_message.strip()

    # 限制长度（最多 50 个字符）
    if len(title) > 50:
        title = title[:50] + "..."

    return title


# ==================== 流式处理 ====================

async def process_query_stream(
    datasource_id: str,
    db_type: str,
    host: str = None,
    port: int = None,
    database: str = None,
    username: str = None,
    password: str = None,
    schema_name: Optional[str] = None,
    tables_info: Optional[List[dict]] = None,
    query: str = None,
    context: dict = None,
    history: list[dict] = None,
    session_id: str = None,  # 新增：使用 session_id 作为 thread_id
) -> AsyncGenerator[dict, None]:
    """流式处理用户查询 - v3.0 生产环境重构

    使用 session_id 作为 thread_id，实现状态持久化。
    使用原生异步流式 API，避免线程与协程混合。
    支持 file 类型数据源（DuckDB）。
    """

    import logging
    logger = logging.getLogger("agent_service")
    step_counter = 0

    def log_step(step_name: str, detail: str = ""):
        nonlocal step_counter
        step_counter += 1
        print(f"[Agent] 步骤 {step_counter}: {step_name} - {detail}")

    log_step("开始处理", f"用户问题：{query[:50]}...")

    # 设置数据库连接上下文（供工具函数使用）
    set_db_context(
        datasource_id=datasource_id,
        db_type=db_type,
        host=host,
        port=port,
        database=database,
        username=username,
        password=password,
        schema_name=schema_name,
        tables_info=tables_info,
    )

    # 发送思考事件
    yield ThinkingEvent(content="正在分析您的问题...").to_dict()

    # 获取 Agent 实例
    agent = await get_agent()
    log_step("获取 Agent 实例", "成功")

    # 使用 session_id 作为 thread_id（如果未提供则生成新的）
    thread_id = session_id or str(uuid.uuid4())
    config = {
        "configurable": {
            "thread_id": thread_id,
        },
    }
    log_step("创建会话", f"thread_id: {thread_id[:8]}... (session_id: {session_id})")

    # 构建消息
    messages = []
    for msg in history[-10:]:
        role = msg.get("role", "user")
        blocks = msg.get("blocks", [])
        text_content = ""
        for block in blocks:
            if block.get("type") == "text":
                text_content = block.get("content", "")
                break
        if not text_content:
            text_content = msg.get("content", "")
        if text_content:
            messages.append({"role": role, "content": text_content})

    # 添加数据源上下文
    messages.append({
        "role": "system",
        "content": f"当前数据源 ID: {datasource_id}。调用工具时使用此 ID。",
    })

    # 添加当前问题
    messages.append({"role": "user", "content": query})
    log_step("构建消息", f"历史消息数：{len(messages) - 2}")

    # 最终结果
    final_result = {
        "content": "",
        "sql": None,
        "error": None,
    }

    try:
        # 使用原生异步流式 API
        async for chunk in agent.astream({"messages": messages}, config=config):
            event = _parse_agent_chunk(chunk)

            if event:
                # 日志打印
                event_type = type(event).__name__
                if isinstance(event, ToolCallEvent):
                    log_step("工具调用", f"{event.tool}")
                elif isinstance(event, ToolResultEvent):
                    log_step("工具结果", f"{event.tool} - {len(event.output or '')} 字符")
                elif isinstance(event, MessageEvent):
                    log_step("生成回复", f"{len(event.content or '')} 字符")
                elif isinstance(event, SQLGenerationEvent):
                    log_step("SQL 生成", event.sql[:50] if event.sql else "")
                elif isinstance(event, SQLExecutionEvent):
                    log_step("SQL 执行", f"状态：{event.status}")

                # 收集消息内容
                if isinstance(event, MessageEvent) and event.content:
                    final_result["content"] = event.content

                # 收集 SQL（用于保存到消息记录）
                if isinstance(event, ToolCallEvent) and event.tool == "run_sql":
                    sql = event.input.get("sql", "") if event.input else ""
                    if sql:
                        final_result["sql"] = sql

                yield event.to_dict()

        log_step("处理完成", f"总步骤数：{step_counter}")
        # 发送完成事件
        yield {"type": "done", "message": "处理完成", "data": final_result}

    except Exception as e:
        log_step("处理出错", str(e))
        yield ErrorEvent(message=f"处理出错：{str(e)}").to_dict()
        yield DoneEvent(message="处理结束").to_dict()



def _parse_agent_chunk(chunk: dict) -> Optional[SSEEvent]:
    """解析 Agent 输出的 chunk，转换为 SSE 事件"""

    # DeepAgents 的 chunk 是嵌套结构
    for key, value in chunk.items():
        if isinstance(value, dict) and "messages" in value:
            messages = value["messages"]

            # 处理 LangGraph 的 Overwrite 对象
            if hasattr(messages, "value"):
                messages = messages.value
            elif not isinstance(messages, (list, tuple)):
                try:
                    messages = list(messages) if messages else []
                except:
                    messages = []

            if messages:
                last_msg = messages[-1]

                # 处理 LangChain 消息对象
                if hasattr(last_msg, "content"):
                    content = last_msg.content

                    # 检查是否有工具调用
                    tool_calls = getattr(last_msg, "tool_calls", None)
                    if tool_calls:
                        tc = tool_calls[-1]
                        tool_name = tc.get("name", "unknown") if isinstance(tc, dict) else getattr(tc, "name", "unknown")
                        if isinstance(tc, dict):
                            return ToolCallEvent(
                                name=tool_name,
                                tool=tool_name,
                                input=tc.get("args", {}),
                                id=tc.get("id", str(uuid.uuid4()))
                            )
                        else:
                            return ToolCallEvent(
                                name=tool_name,
                                tool=tool_name,
                                input=getattr(tc, "args", {}) or {},
                                id=getattr(tc, "id", str(uuid.uuid4()))
                            )

                    # 检查是否是工具结果消息
                    msg_type = getattr(last_msg, "type", None)
                    msg_name = getattr(last_msg, "name", None)
                    if msg_type == "tool" and msg_name:
                        return ToolResultEvent(
                            name=msg_name,
                            tool=msg_name,
                            id=str(uuid.uuid4()),
                            output=json.dumps(content) if isinstance(content, dict) else str(content)[:2000] if content else ""
                        )

                    # 过滤掉 human/user 消息
                    if msg_type == "human":
                        return None

                    # 普通 AI 消息
                    # content 可能是字符串或字符串列表
                    if content:
                        if isinstance(content, list):
                            # 流式内容可能是字符串列表
                            content_str = "".join(str(c) if isinstance(c, str) else "" for c in content)
                        else:
                            content_str = str(content)

                        if content_str.strip():
                            return MessageEvent(content=content_str)

                # 处理字典格式的消息
                elif isinstance(last_msg, dict):
                    msg_content = last_msg.get("content", "")
                    tool_calls = last_msg.get("tool_calls", [])
                    msg_type = last_msg.get("type", "")
                    msg_name = last_msg.get("name", "")

                    if tool_calls:
                        tc = tool_calls[-1]
                        tool_name = tc.get("name", msg_name or "unknown")
                        return ToolCallEvent(
                            name=tool_name,
                            tool=tool_name,
                            input=tc.get("args", {}),
                            id=tc.get("id", str(uuid.uuid4()))
                        )

                    if msg_type == "tool":
                        return ToolResultEvent(
                            name=msg_name,
                            tool=msg_name,
                            id=str(uuid.uuid4()),
                            output=msg_content if isinstance(msg_content, dict) else str(msg_content)[:2000] if msg_content else ""
                        )

                    if msg_type == "human":
                        return None

                    # 处理消息内容（可能是字符串或列表）
                    if msg_content:
                        if isinstance(msg_content, list):
                            content_str = "".join(str(c) if isinstance(c, str) else "" for c in msg_content)
                        else:
                            content_str = str(msg_content)

                        if content_str.strip():
                            return MessageEvent(content=content_str)

    # 处理 todos 事件
    if "todos" in chunk:
        return PlanEvent(todos=chunk["todos"])

    return None


# ==================== 兼容性接口 ====================

async def process_query(
    datasource_id: str,
    db_type: str,
    host: str,
    port: int,
    database: str,
    username: str,
    password: str,
    schema_name: Optional[str],
    query: str,
    context: dict,
    history: list[dict],
    session_id: str = None,
) -> dict:
    """非流式处理用户查询（兼容性接口）"""
    result = None

    async for event in process_query_stream(
        datasource_id, db_type, host, port, database,
        username, password, schema_name, query, context, history, session_id
    ):
        if event.get("type") == "message":
            result = {"content": event.get("content")}
        elif event.get("type") == "sql_generation":
            if result is None:
                result = {}
            result["sql"] = event.get("sql")

    return result or {"content": "处理完成", "error": None}
