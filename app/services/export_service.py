"""数据导出服务"""

import io
import csv
from datetime import datetime
from typing import List, Any, Optional

from app.models import ExportFormat, ExportLimit


def export_to_csv(
    columns: List[str],
    rows: List[List[Any]],
    filename: Optional[str] = None,
) -> tuple[bytes, str]:
    """
    导出数据为 CSV 格式。

    Args:
        columns: 列名列表
        rows: 数据行列表
        filename: 文件名（不含扩展名）

    Returns:
        tuple[bytes, str]: (文件内容, 文件名)
    """
    # 检查限制
    if len(rows) > ExportLimit.MAX_ROWS:
        raise ValueError(f"数据行数超过限制（最大 {ExportLimit.MAX_ROWS} 行）")
    if len(columns) > ExportLimit.MAX_COLUMNS:
        raise ValueError(f"数据列数超过限制（最大 {ExportLimit.MAX_COLUMNS} 列）")

    # 生成 CSV
    output = io.StringIO()
    writer = csv.writer(output, quoting=csv.QUOTE_MINIMAL)

    # 写入表头
    writer.writerow(columns)

    # 写入数据
    for row in rows:
        # 处理特殊值
        processed_row = []
        for value in row:
            if value is None:
                processed_row.append("")
            elif isinstance(value, datetime):
                processed_row.append(value.strftime("%Y-%m-%d %H:%M:%S"))
            elif isinstance(value, bytes):
                processed_row.append(value.decode("utf-8", errors="replace"))
            else:
                processed_row.append(str(value))
        writer.writerow(processed_row)

    # 生成文件名
    if not filename:
        filename = f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}"

    return output.getvalue().encode("utf-8-sig"), f"{filename}.csv"


def export_to_excel(
    columns: List[str],
    rows: List[List[Any]],
    filename: Optional[str] = None,
    title: Optional[str] = None,
) -> tuple[bytes, str]:
    """
    导出数据为 Excel 格式。

    Args:
        columns: 列名列表
        rows: 数据行列表
        filename: 文件名（不含扩展名）
        title: 工作表标题

    Returns:
        tuple[bytes, str]: (文件内容, 文件名)
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("请安装 openpyxl 库: pip install openpyxl")

    # 检查限制
    if len(rows) > ExportLimit.MAX_ROWS:
        raise ValueError(f"数据行数超过限制（最大 {ExportLimit.MAX_ROWS} 行）")
    if len(columns) > ExportLimit.MAX_COLUMNS:
        raise ValueError(f"数据列数超过限制（最大 {ExportLimit.MAX_COLUMNS} 列）")

    # 创建工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title or "数据"

    # 样式定义
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # 写入表头
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 写入数据
    for row_idx, row in enumerate(rows, 2):
        for col_idx, value in enumerate(row, 1):
            # 处理特殊值
            if value is None:
                processed_value = ""
            elif isinstance(value, datetime):
                processed_value = value.strftime("%Y-%m-%d %H:%M:%S")
            elif isinstance(value, bytes):
                processed_value = value.decode("utf-8", errors="replace")
            else:
                processed_value = value

            cell = ws.cell(row=row_idx, column=col_idx, value=processed_value)
            cell.border = thin_border

            # 数值格式化
            if isinstance(processed_value, float):
                cell.number_format = "#,##0.00"

    # 自动调整列宽
    for col_idx, col_name in enumerate(columns, 1):
        # 计算最大宽度
        max_length = len(str(col_name))
        for row in rows:
            if col_idx <= len(row):
                cell_value = row[col_idx - 1]
                if cell_value is not None:
                    max_length = max(max_length, len(str(cell_value)))

        # 设置列宽（最大50）
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

    # 冻结首行
    ws.freeze_panes = "A2"

    # 添加自动筛选
    if columns:
        last_col = get_column_letter(len(columns))
        ws.auto_filter.ref = f"A1:{last_col}{len(rows) + 1}"

    # 保存到字节流
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # 生成文件名
    if not filename:
        filename = f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}"

    return output.getvalue(), f"{filename}.xlsx"


def export_data(
    columns: List[str],
    rows: List[List[Any]],
    format: ExportFormat,
    filename: Optional[str] = None,
    title: Optional[str] = None,
) -> tuple[bytes, str, str]:
    """
    导出数据。

    Args:
        columns: 列名列表
        rows: 数据行列表
        format: 导出格式
        filename: 文件名（不含扩展名）
        title: 标题

    Returns:
        tuple[bytes, str, str]: (文件内容, 文件名, MIME类型)
    """
    if format == ExportFormat.CSV:
        content, filename = export_to_csv(columns, rows, filename)
        mime_type = "text/csv"
    elif format == ExportFormat.EXCEL:
        content, filename = export_to_excel(columns, rows, filename, title)
        mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    else:
        raise ValueError(f"不支持的导出格式: {format}")

    return content, filename, mime_type